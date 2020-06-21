#Code to access rally userboards and get the status and update the excel sheets with values against the respective states.

#Step 1 : Connect to rally using pyral
from pyral import *
from pyral import Rally

SERVER = 'SERVER'
USER = 'USER'
PASSWORD = 'PASSWORD'
WORKSPACE = 'WORKSPACE'
PROJECT = 'PROJECT'

rally = Rally(SERVER, USER, PASSWORD, workspace=WORKSPACE)

project_req = rally.get('Project', fetch=True, query='Name = "%s"' % (PROJECT))
project = ''

user_stories = rally.get('HierarchicalRequirement', fetch=True, query='Project = %s' % (project.ref))

#for user_story in user_stories:
        
#Step 2 : Access all userstories that has names 'Disability-Policy Equipment' or 'Enahnaced-Life policy Equipment'
#(100) and get their status.

#Step 3 : Assign status values that has to be updated on excel sheet.
#Values are assigned based on the lane the user story belongs to.
lane = ''
Status = ''
State = ''
if lane=='New' :
    FinalValue = '0'
elif lane=='Issue Requirements' and  Status=='ready':
    FinalValue = '25'
elif lane=='Issue Requirements' and  Status!='ready':   
    FinalValue = '0'
elif lane=='IT Requirements' :
    FinalValue = '25'
elif lane=='PD Rules Review' and  Status=='ready':
    FinalValue = '50'
elif lane=='PD Rules Review' and  Status!='ready':
    FinalValue = '25'
elif lane=='PDA Development' :
    FinalValue = '50'
elif lane=='Booklets Development' and  Status=='ready':
    FinalValue = '75'
elif lane=='Booklets Development' and  Status!='ready':
    FinalValue = '50'
elif lane=='QA validation':
    FinalValue = '75'
elif lane=='PD validation' and  Status=='ready':
    FinalValue = '100'
elif lane=='PD validation' and  Status!='ready':
    FinalValue = '75'
elif lane=='UAT':
    FinalValue = '100'
elif lane=='Ready for production':
    FinalValue = '100'
elif lane=='Released to production':
    FinalValue = 'Done'

#Step 4 : Connect to excel sheet and update the values.

import openpyxl
from openpyxl import Workbook,load_workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "State"
sheet["B1"] = "DI"
sheet["C1"] = "Life"
sheet["A2"] = "AL"
sheet["A3"] = "CT"
sheet["A4"] = "NY"
sheet["A5"] = "OH"

workbook.save(filename="hello_world.xlsx")

workbook = load_workbook(filename="hello_world.xlsx")
#print(workbook.sheetnames)

sheet = workbook.active
print(sheet)
print(sheet.max_row)

        
PRICE_UPDATES = {'AL': 100,'CT': 75, 'OH': 25,'NY':0}
        
        # Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row): # skip the first row
   statename = sheet.cell(row=rowNum, column=1).value
   if statename in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[statename]

workbook.save('hello_world.xlsx')
